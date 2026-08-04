import io
import pandas as pd
from typing import Dict, Any, List
from datetime import date
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from repositories.transaction_repo import TransactionRepository
from services.analytics_service import AnalyticsService
from utils.pdf_generator import generate_financial_pdf_report
from utils.formatters import format_currency, format_date

class ReportExportService:
    def __init__(self, db: Session):
        self.db = db
        self.trans_repo = TransactionRepository(db)
        self.analytics = AnalyticsService(db)

    def get_transactions_dataframe(self, start_date: date, end_date: date, **kwargs) -> pd.DataFrame:
        transactions = self.trans_repo.get_filtered(start_date=start_date, end_date=end_date, **kwargs)
        data = []
        for t in transactions:
            data.append({
                "ID": t.id,
                "data": t.transaction_date,
                "descricao": t.description,
                "tipo": t.type,
                "valor": t.amount,
                "forma_pagamento": t.payment_method,
                "status": t.status,
                "categoria": t.category.name if t.category else "",
                "subcategoria": t.subcategory.name if t.subcategory else "",
                "conta": t.account.name if t.account else "",
                "cartao": t.card.name if t.card else "",
                "proprietario": t.owner.name if t.owner else "",
                "tags": t.tags,
                "observacoes": t.notes
            })
        return pd.DataFrame(data)

    def export_to_excel(self, start_date: date, end_date: date) -> bytes:
        """Gera um arquivo Excel multi-aba com formatação profissional (OpenPyXL)."""
        df_trans = self.get_transactions_dataframe(start_date, end_date)
        kpis = self.analytics.get_summary_kpis(start_date, end_date)

        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        
        # Aba 1: Resumo Executivo
        ws_kpi = wb.active
        ws_kpi.title = "Resumo Executivo"
        
        # Estilos
        title_font = Font(name="Arial", size=16, bold=True, color="1E293B")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell_font = Font(name="Arial", size=10)
        bold_font = Font(name="Arial", size=10, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws_kpi.append(["RELATÓRIO FINANCEIRO EXECUTIVO"])
        ws_kpi.cell(row=1, column=1).font = title_font
        ws_kpi.append([f"Período: {format_date(start_date)} a {format_date(end_date)}"])
        ws_kpi.append([])

        # KPI Table
        ws_kpi.append(["Indicador", "Valor"])
        for cell in ws_kpi[4]:
            cell.font = header_font
            cell.fill = header_fill

        kpi_rows = [
            ("Receitas Totais", kpis['receitas']),
            ("Despesas Totais", kpis['despesas']),
            ("Saldo do Período", kpis['saldo']),
            ("Taxa de Economia (%)", f"{kpis['taxa_economia']:.1f}%"),
            ("Média Diária de Gastos", kpis['media_diaria_gasto']),
            ("Ticket Médio de Despesas", kpis['ticket_medio_despesa']),
            ("Total de Movimentações", kpis['qtd_movimentacoes'])
        ]

        for label, val in kpi_rows:
            ws_kpi.append([label, val])

        for row in ws_kpi.iter_rows(min_row=4, max_row=11, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                cell.font = cell_font

        # Aba 2: Movimentações
        ws_data = wb.create_sheet(title="Movimentações")
        if not df_trans.empty:
            headers = list(df_trans.columns)
            ws_data.append(headers)
            for cell in ws_data[1]:
                cell.font = header_font
                cell.fill = header_fill

            for _, r in df_trans.iterrows():
                row_vals = [r[col] for col in headers]
                ws_data.append(row_vals)

            for row in ws_data.iter_rows(min_row=1, max_row=len(df_trans)+1, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border

        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes

    def export_to_pdf(self, start_date: date, end_date: date) -> bytes:
        df_trans = self.get_transactions_dataframe(start_date, end_date)
        kpis = self.analytics.get_summary_kpis(start_date, end_date)
        period_str = f"{format_date(start_date)} a {format_date(end_date)}"
        return generate_financial_pdf_report("Relatório Financeiro Pessoal", period_str, kpis, df_trans)
